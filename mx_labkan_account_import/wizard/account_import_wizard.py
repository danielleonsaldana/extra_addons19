# -*- coding: utf-8 -*-
import base64
import io
import logging

from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
#  Tipos de cuenta válidos en Odoo 17
# ──────────────────────────────────────────────
ACCOUNT_TYPE_MAP = {
    # Español
    'activo corriente': 'asset_current',
    'activo no corriente': 'asset_non_current',
    'activo fijo': 'asset_fixed',
    'prepagos': 'asset_prepayments',
    'cuentas por cobrar': 'asset_receivable',
    'banco y efectivo': 'asset_cash',
    'cuentas por pagar': 'liability_payable',
    'tarjeta de crédito': 'liability_credit_card',
    'pasivo corriente': 'liability_current',
    'pasivo no corriente': 'liability_non_current',
    'capital': 'equity',
    'utilidades del ejercicio': 'equity_unaffected',
    'ingresos': 'income',
    'otros ingresos': 'income_other',
    'gastos': 'expense',
    'otros gastos': 'expense_other',
    'depreciación': 'expense_depreciation',
    'costo de ventas': 'expense_direct_cost',
    'fuera de balance': 'off_balance',
    # Inglés (técnico)
    'asset_receivable': 'asset_receivable',
    'asset_cash': 'asset_cash',
    'asset_current': 'asset_current',
    'asset_non_current': 'asset_non_current',
    'asset_prepayments': 'asset_prepayments',
    'asset_fixed': 'asset_fixed',
    'liability_payable': 'liability_payable',
    'liability_credit_card': 'liability_credit_card',
    'liability_current': 'liability_current',
    'liability_non_current': 'liability_non_current',
    'equity': 'equity',
    'equity_unaffected': 'equity_unaffected',
    'income': 'income',
    'income_other': 'income_other',
    'expense': 'expense',
    'expense_other': 'expense_other',
    'expense_depreciation': 'expense_depreciation',
    'expense_direct_cost': 'expense_direct_cost',
    'off_balance': 'off_balance',
}

VALID_ACCOUNT_TYPES = list(set(ACCOUNT_TYPE_MAP.values()))


class AccountImportWizard(models.TransientModel):
    _name = 'mx.account.import.wizard'
    _description = 'Alta Masiva de Cuentas Contables'

    # ── Archivo ──────────────────────────────────
    import_file = fields.Binary(
        string='Archivo Excel/CSV',
        required=True,
        attachment=False,
    )
    import_filename = fields.Char(string='Nombre de archivo')
    file_type = fields.Selection([
        ('excel', 'Excel (.xlsx / .xls)'),
        ('csv', 'CSV'),
    ], string='Tipo de archivo', default='excel', required=True)

    # ── Opciones ─────────────────────────────────
    company_id = fields.Many2one(
        'res.company',
        string='Compañía',
        required=True,
        default=lambda self: self.env.company,
    )
    sheet_name = fields.Char(
        string='Nombre de hoja (Excel)',
        default='Cuentas',
        help='Nombre exacto de la hoja en el archivo Excel. Si se deja vacío se usa la primera hoja.',
    )
    header_row = fields.Integer(
        string='Fila de encabezados',
        default=1,
        help='Número de fila donde están los encabezados de columna (1 = primera fila).',
    )
    update_existing = fields.Boolean(
        string='Actualizar cuentas existentes',
        default=False,
        help='Si está activo, las cuentas cuyo código ya exista serán actualizadas.',
    )

    # ── Resultados ───────────────────────────────
    state = fields.Selection([
        ('draft', 'Pendiente'),
        ('done', 'Importado'),
        ('error', 'Con errores'),
    ], default='draft')
    result_message = fields.Text(string='Resultado', readonly=True)
    error_log = fields.Text(string='Errores detallados', readonly=True)

    # ── Plantilla ────────────────────────────────
    def action_download_template(self):
        """Genera y descarga un archivo Excel de plantilla."""
        try:
            import openpyxl
        except ImportError:
            raise UserError(_('openpyxl no está instalado. Instala con: pip install openpyxl'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cuentas'

        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='2E7D32')  # Verde Labkan

        headers = [
            ('codigo', 'Código *', 'Ej: 101.01.001'),
            ('nombre', 'Nombre *', 'Ej: Caja General'),
            ('tipo', 'Tipo de cuenta *', 'Ej: asset_current'),
            ('reconciliable', 'Reconciliable', 'Sí / No'),
            ('moneda', 'Moneda', 'Ej: MXN (opcional)'),
            ('etiquetas', 'Etiquetas', 'Separadas por | (opcional)'),
            ('notas', 'Notas internas', 'Texto libre (opcional)'),
            ('activo', 'Activo', 'Sí / No (default: Sí)'),
        ]

        for col, (field, label, example) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.cell(row=2, column=col, value=example)

        # Fila de ejemplo
        examples = [
            '101.01.001', 'Caja General MXN', 'asset_cash',
            'No', 'MXN', '', '', 'Sí',
        ]
        for col, val in enumerate(examples, 1):
            ws.cell(row=3, column=col, value=val)

        # Hoja de referencia de tipos
        ws2 = wb.create_sheet('Tipos de cuenta')
        ws2.append(['Valor técnico', 'Descripción'])
        type_desc = [
            ('asset_receivable', 'Cuentas por cobrar'),
            ('asset_cash', 'Banco y Efectivo'),
            ('asset_current', 'Activo Corriente'),
            ('asset_non_current', 'Activo No Corriente'),
            ('asset_prepayments', 'Prepagos'),
            ('asset_fixed', 'Activo Fijo'),
            ('liability_payable', 'Cuentas por pagar'),
            ('liability_credit_card', 'Tarjeta de Crédito'),
            ('liability_current', 'Pasivo Corriente'),
            ('liability_non_current', 'Pasivo No Corriente'),
            ('equity', 'Capital'),
            ('equity_unaffected', 'Utilidades del ejercicio'),
            ('income', 'Ingresos'),
            ('income_other', 'Otros ingresos'),
            ('expense', 'Gastos'),
            ('expense_other', 'Otros gastos'),
            ('expense_depreciation', 'Depreciación'),
            ('expense_direct_cost', 'Costo de Ventas'),
            ('off_balance', 'Fuera de balance'),
        ]
        for row in type_desc:
            ws2.append(row)

        # Ajustar anchos
        for ws_ in [ws, ws2]:
            for col in ws_.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws_.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_data = base64.b64encode(buffer.read())

        attachment = self.env['ir.attachment'].create({
            'name': 'plantilla_alta_cuentas.xlsx',
            'datas': file_data,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ── Importar ─────────────────────────────────
    def action_import(self):
        self.ensure_one()
        if not self.import_file:
            raise UserError(_('Por favor selecciona un archivo para importar.'))

        rows = self._parse_file()
        created, updated, errors = self._process_rows(rows)

        msg_parts = [
            f'✅ Cuentas creadas: {created}',
            f'🔄 Cuentas actualizadas: {updated}',
            f'❌ Errores: {len(errors)}',
        ]
        self.result_message = '\n'.join(msg_parts)
        self.error_log = '\n'.join(errors) if errors else 'Sin errores.'
        self.state = 'error' if errors and not (created + updated) else 'done'

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    # ── Parser ───────────────────────────────────
    def _parse_file(self):
        """Devuelve lista de dicts con los datos de cada fila."""
        file_data = base64.b64decode(self.import_file)
        if self.file_type == 'excel':
            return self._parse_excel(file_data)
        else:
            return self._parse_csv(file_data)

    def _parse_excel(self, file_data):
        try:
            import openpyxl
        except ImportError:
            raise UserError(_('openpyxl no está instalado.'))

        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)

        # Selección de hoja
        if self.sheet_name and self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_('El archivo Excel está vacío.'))

        header_idx = self.header_row - 1
        if header_idx >= len(rows):
            raise UserError(_('La fila de encabezados indicada (%s) supera el número de filas del archivo.') % self.header_row)

        headers = [str(h).strip().lower() if h is not None else '' for h in rows[header_idx]]
        data_rows = rows[header_idx + 1:]

        return self._rows_to_dicts(headers, data_rows)

    def _parse_csv(self, file_data):
        import csv
        content = file_data.decode('utf-8-sig').splitlines()
        reader = csv.reader(content)
        rows = list(reader)
        if not rows:
            raise UserError(_('El archivo CSV está vacío.'))

        header_idx = self.header_row - 1
        headers = [h.strip().lower() for h in rows[header_idx]]
        data_rows = [row for row in rows[header_idx + 1:] if any(r.strip() for r in row)]

        return self._rows_to_dicts(headers, data_rows)

    def _rows_to_dicts(self, headers, data_rows):
        """Normaliza encabezados y convierte filas a lista de dicts."""
        # Mapa de alias de columnas (flexibilidad de nombres)
        col_alias = {
            'código': 'codigo', 'codigo': 'codigo', 'code': 'codigo',
            'nombre': 'nombre', 'name': 'nombre', 'cuenta': 'nombre',
            'tipo': 'tipo', 'type': 'tipo', 'tipo de cuenta': 'tipo',
            'account type': 'tipo',
            'reconciliable': 'reconciliable', 'reconcile': 'reconciliable',
            'moneda': 'moneda', 'currency': 'moneda',
            'etiquetas': 'etiquetas', 'tags': 'etiquetas',
            'notas': 'notas', 'notes': 'notas', 'nota': 'notas',
            'activo': 'activo', 'active': 'activo',
        }
        normalized_headers = [col_alias.get(h, h) for h in headers]

        result = []
        for row in data_rows:
            row_dict = {}
            for i, header in enumerate(normalized_headers):
                val = row[i] if i < len(row) else None
                row_dict[header] = str(val).strip() if val is not None else ''
            # Ignorar filas completamente vacías
            if any(v for v in row_dict.values()):
                result.append(row_dict)
        return result

    # ── Procesamiento ─────────────────────────────
    def _process_rows(self, rows):
        created = updated = 0
        errors = []
        AccountAccount = self.env['account.account']

        for idx, row in enumerate(rows, start=self.header_row + 1):
            try:
                vals = self._prepare_account_vals(row, idx)
                if vals is None:
                    continue  # fila vacía

                code = vals.get('code')
                existing = AccountAccount.with_context(
                    allowed_company_ids=[self.company_id.id]
                ).search([
                    ('code', '=', code),
                    ('company_ids', 'in', [self.company_id.id]),
                ], limit=1)

                if existing:
                    if self.update_existing:
                        existing.write(vals)
                        updated += 1
                    else:
                        errors.append(
                            f'Fila {idx}: Código {code} ya existe (activa la opción "Actualizar" para sobreescribir).'
                        )
                else:
                    vals['company_ids'] = [fields.Command.set([self.company_id.id])]
                    AccountAccount.create(vals)
                    created += 1

            except (UserError, ValidationError) as e:
                errors.append(f'Fila {idx}: {e.args[0]}')
            except Exception as e:
                _logger.exception('Error en fila %s: %s', idx, e)
                errors.append(f'Fila {idx}: Error inesperado — {e}')

        return created, updated, errors

    def _prepare_account_vals(self, row, row_num):
        """Convierte un dict de fila a vals para account.account."""
        code = row.get('codigo', '').strip()
        name = row.get('nombre', '').strip()
        account_type_raw = row.get('tipo', '').strip().lower()

        # Validaciones obligatorias
        if not code and not name:
            return None  # fila vacía, ignorar
        if not code:
            raise UserError(_(f'Columna "Código" vacía.'))
        if not name:
            raise UserError(_(f'Columna "Nombre" vacía para código {code}.'))
        if not account_type_raw:
            raise UserError(_(f'Columna "Tipo de cuenta" vacía para {code}.'))

        account_type = ACCOUNT_TYPE_MAP.get(account_type_raw)
        if not account_type:
            raise UserError(
                _(f'Tipo de cuenta inválido: "{account_type_raw}". '
                  f'Valores permitidos: {", ".join(VALID_ACCOUNT_TYPES)}')
            )

        # Reconciliable
        reconcile_raw = row.get('reconciliable', '').lower()
        reconcile = reconcile_raw in ('sí', 'si', 'yes', 'true', '1', 'x')
        # Receivable/Payable DEBEN ser reconciliables
        if account_type in ('asset_receivable', 'liability_payable'):
            reconcile = True

        # Moneda
        currency_id = False
        currency_raw = row.get('moneda', '').strip().upper()
        if currency_raw:
            currency = self.env['res.currency'].search([('name', '=', currency_raw)], limit=1)
            if not currency:
                raise UserError(_(f'Moneda no encontrada: "{currency_raw}" para cuenta {code}.'))
            currency_id = currency.id

        # Activo
        active_raw = row.get('activo', '').lower()
        active = active_raw not in ('no', 'false', '0', 'inactivo') if active_raw else True

        # Notas
        note = row.get('notas', '').strip() or False

        vals = {
            'name': name,
            'code': code,
            'account_type': account_type,
            'reconcile': reconcile,
            'active': active,
        }
        if currency_id:
            vals['currency_id'] = currency_id
        if note:
            vals['note'] = note

        return vals
